"""
Enhanced Specialized Agents
XAgent-inspired specialized agents with advanced capabilities for
document processing, price analysis, trend prediction, and customer insights.
"""

import asyncio
import json
import logging
import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
import hashlib
import re
import time
import statistics
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from abc import ABC, abstractmethod

# Machine learning and analytics
from sklearn.ensemble import RandomForestRegressor, IsolationForest
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfVectorizer
import joblib

# External integrations
import requests
import aiohttp
import pdfplumber
import openpyxl
from PIL import Image
import pytesseract

# Import agent framework
from multi_agent_orchestrator import BaseAgent, AgentTask, AgentCapability, AgentStatus
from multi_agent_system.protocols.agent_communication import (
    MessageRouter, TaskDelegator, AgentMessage, TaskRequest, TaskResponse,
    MessageType, Priority
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ProcessingMetrics:
    """Metrics for agent performance"""
    total_processed: int = 0
    success_rate: float = 0.0
    average_processing_time: float = 0.0
    quality_score: float = 0.0
    error_count: int = 0
    last_processed: Optional[datetime] = None


class EnhancedBaseAgent(BaseAgent):
    """Enhanced base agent with advanced capabilities"""

    def __init__(self, agent_id: str, name: str, orchestrator):
        super().__init__(agent_id, name, orchestrator)
        self.message_router: Optional[MessageRouter] = None
        self.task_delegator: Optional[TaskDelegator] = None
        self.metrics = ProcessingMetrics()
        self.capabilities: Set[str] = set()
        self.performance_history: List[Dict[str, Any]] = []
        self.resource_usage = defaultdict(float)
        self.active_tasks: Set[str] = set()

    async def initialize(self):
        """Initialize enhanced agent"""
        await super().initialize()

        # Initialize communication
        self.message_router = MessageRouter(self.id)
        await self.message_router.start()

        self.task_delegator = TaskDelegator(self.message_router)

        # Register message handlers
        self.message_router.register_handler(MessageType.TASK_DELEGATION, self._handle_task_delegation)
        self.message_router.register_handler(MessageType.COLLABORATION_REQUEST, self._handle_collaboration_request)

        # Update capabilities in delegator
        self.task_delegator.update_agent_capabilities(self.id, self.get_capabilities())

        logger.info(f"Enhanced agent {self.id} initialized with {len(self.capabilities)} capabilities")

    async def _handle_task_delegation(self, message: AgentMessage):
        """Handle task delegation"""
        try:
            task_data = message.payload.get('task_request', {})
            task = TaskRequest.from_dict(task_data)

            # Execute task
            result = await self.execute_delegated_task(task)

            # Send response
            response = AgentMessage(
                sender_id=self.id,
                receiver_id=message.sender_id,
                message_type=MessageType.TASK_RESPONSE,
                reply_to=message.message_id,
                correlation_id=message.correlation_id,
                payload={
                    'task_id': task.task_id,
                    'response': result.to_dict() if isinstance(result, TaskResponse) else result
                }
            )

            await self.message_router.send_message(response)

        except Exception as e:
            logger.error(f"Error handling task delegation: {e}")
            error_response = AgentMessage(
                sender_id=self.id,
                receiver_id=message.sender_id,
                message_type=MessageType.ERROR,
                reply_to=message.message_id,
                payload={'error': str(e)}
            )
            await self.message_router.send_message(error_response)

    async def _handle_collaboration_request(self, message: AgentMessage):
        """Handle collaboration requests"""
        try:
            collaboration_data = message.payload
            role = collaboration_data.get('role', 'supporting')

            if role == 'primary':
                # Handle as primary agent
                await self._handle_primary_collaboration(collaboration_data)
            else:
                # Handle as supporting agent
                await self._handle_supporting_collaboration(collaboration_data)

        except Exception as e:
            logger.error(f"Error handling collaboration request: {e}")

    async def execute_delegated_task(self, task: TaskRequest) -> TaskResponse:
        """Execute a delegated task"""
        start_time = time.time()

        try:
            # Add to active tasks
            self.active_tasks.add(task.task_id)

            # Update resource usage
            self._update_resource_usage('cpu', 0.1)
            self._update_resource_usage('memory', 0.05)

            # Execute the task
            result = await self._execute_task_logic(task)

            # Calculate metrics
            execution_time = time.time() - start_time

            # Update performance metrics
            self._update_metrics(True, execution_time)

            # Create response
            response = TaskResponse(
                task_id=task.task_id,
                success=True,
                result=result,
                execution_time=execution_time,
                resources_used=dict(self.resource_usage)
            )

            return response

        except Exception as e:
            execution_time = time.time() - start_time
            self._update_metrics(False, execution_time)

            return TaskResponse(
                task_id=task.task_id,
                success=False,
                error_message=str(e),
                execution_time=execution_time
            )

        finally:
            # Remove from active tasks
            self.active_tasks.discard(task.task_id)
            # Reset resource usage
            self.resource_usage.clear()

    async def _execute_task_logic(self, task: TaskRequest) -> Any:
        """Override in subclasses"""
        raise NotImplementedError

    def _update_resource_usage(self, resource_type: str, amount: float):
        """Update resource usage"""
        self.resource_usage[resource_type] += amount

    def _update_metrics(self, success: bool, execution_time: float):
        """Update performance metrics"""
        self.metrics.total_processed += 1

        if success:
            # Update processing time
            total_time = self.metrics.average_processing_time * (self.metrics.total_processed - 1)
            self.metrics.average_processing_time = (total_time + execution_time) / self.metrics.total_processed

            # Update success rate
            self.metrics.success_rate = (self.metrics.success_rate * (self.metrics.total_processed - 1) + 1.0) / self.metrics.total_processed
        else:
            self.metrics.error_count += 1
            self.metrics.success_rate = (self.metrics.success_rate * (self.metrics.total_processed - 1)) / self.metrics.total_processed

        self.metrics.last_processed = datetime.now()

    def get_performance_metrics(self) -> Dict[str, Any]:
        """Get performance metrics"""
        return {
            'agent_id': self.id,
            'metrics': self.metrics.__dict__,
            'active_tasks': len(self.active_tasks),
            'capabilities': list(self.capabilities),
            'performance_history': self.performance_history[-10:]  # Last 10 entries
        }


class DocumentProcessorAgent(EnhancedBaseAgent):
    """Advanced document processing agent with multi-format support"""

    def __init__(self, orchestrator):
        super().__init__("document_processor", "Document Processor", orchestrator)
        self.capabilities.update({
            "document_processing",
            "pdf_parsing",
            "excel_processing",
            "image_ocr",
            "text_extraction",
            "document_classification",
            "metadata_extraction",
            "format_conversion",
            "multi_language_support"
        })
        self.supported_formats = {'.pdf', '.xlsx', '.xls', '.docx', '.doc', '.txt', '.csv', '.jpg', '.jpeg', '.png', '.tiff'}
        self.processing_models = {}
        self.document_cache = {}

    async def initialize(self):
        """Initialize document processing models"""
        await super().initialize()

        # Initialize ML models
        await self._initialize_models()

        logger.info("Document Processor Agent initialized with advanced models")

    async def _initialize_models(self):
        """Initialize ML models for document processing"""
        try:
            # Document classifier
            self.document_classifier = joblib.load('models/document_classifier.pkl')

            # Text extraction quality evaluator
            self.quality_evaluator = joblib.load('models/quality_evaluator.pkl')

            # Language detector
            self.language_detector = joblib.load('models/language_detector.pkl')

        except FileNotFoundError:
            logger.warning("Pre-trained models not found, using fallback methods")
            self.document_classifier = None
            self.quality_evaluator = None
            self.language_detector = None

    async def _execute_task_logic(self, task: TaskRequest) -> Dict[str, Any]:
        """Execute document processing task"""
        task_type = task.task_type
        parameters = task.parameters

        if task_type == "process_document":
            return await self._process_single_document(parameters)
        elif task_type == "batch_process":
            return await self._batch_process_documents(parameters)
        elif task_type == "extract_metadata":
            return await self._extract_document_metadata(parameters)
        elif task_type == "classify_document":
            return await self._classify_document(parameters)
        elif task_type == "convert_format":
            return await self._convert_document_format(parameters)
        elif task_type == "ocr_process":
            return await self._ocr_process_image(parameters)
        else:
            raise ValueError(f"Unknown task type: {task_type}")

    async def _process_single_document(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Process a single document"""
        file_path = parameters.get('file_path')
        if not file_path or not Path(file_path).exists():
            raise ValueError(f"File not found: {file_path}")

        file_path = Path(file_path)
        file_ext = file_path.suffix.lower()

        if file_ext not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_ext}")

        try:
            # Extract content based on file type
            content, metadata = await self._extract_content(file_path)

            # Analyze document quality
            quality_score = await self._evaluate_quality(content, metadata)

            # Classify document type
            document_type = await self._classify_document_type(content, metadata)

            # Extract structured information
            structured_data = await self._extract_structured_info(content, document_type)

            # Cache processing results
            doc_hash = hashlib.md5(str(file_path).encode()).hexdigest()
            self.document_cache[doc_hash] = {
                'content': content,
                'metadata': metadata,
                'structured_data': structured_data,
                'quality_score': quality_score,
                'document_type': document_type,
                'processed_at': datetime.now().isoformat()
            }

            return {
                'file_path': str(file_path),
                'content': content,
                'metadata': metadata,
                'structured_data': structured_data,
                'quality_score': quality_score,
                'document_type': document_type,
                'processing_stats': {
                    'content_length': len(content),
                    'extraction_confidence': quality_score,
                    'processing_time': time.time()
                }
            }

        except Exception as e:
            logger.error(f"Error processing document {file_path}: {e}")
            raise

    async def _extract_content(self, file_path: Path) -> Tuple[str, Dict[str, Any]]:
        """Extract content and metadata from document"""
        file_ext = file_path.suffix.lower()
        metadata = {'file_type': file_ext, 'file_size': file_path.stat().st_size}

        if file_ext == '.pdf':
            return await self._extract_pdf_content(file_path, metadata)
        elif file_ext in ['.xlsx', '.xls']:
            return await self._extract_excel_content(file_path, metadata)
        elif file_ext in ['.docx', '.doc']:
            return await self._extract_word_content(file_path, metadata)
        elif file_ext in ['.txt', '.csv']:
            return await self._extract_text_content(file_path, metadata)
        elif file_ext in ['.jpg', '.jpeg', '.png', '.tiff']:
            return await self._extract_image_content(file_path, metadata)
        else:
            raise ValueError(f"Unsupported file type: {file_ext}")

    async def _extract_pdf_content(self, file_path: Path, metadata: Dict[str, Any]) -> Tuple[str, Dict[str, Any]]:
        """Extract content from PDF file"""
        content = []

        try:
            with pdfplumber.open(file_path) as pdf:
                metadata.update({
                    'page_count': len(pdf.pages),
                    'title': pdf.metadata.get('Title', ''),
                    'author': pdf.metadata.get('Author', ''),
                    'creator': pdf.metadata.get('Creator', ''),
                    'producer': pdf.metadata.get('Producer', ''),
                    'creation_date': pdf.metadata.get('CreationDate', '')
                })

                for page_num, page in enumerate(pdf.pages):
                    page_text = page.extract_text()
                    if page_text:
                        content.append({
                            'page': page_num + 1,
                            'text': page_text,
                            'bbox': page.bbox
                        })

            # Extract tables if available
            tables = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_tables = page.extract_tables()
                    if page_tables:
                        tables.extend(page_tables)

            metadata['tables_found'] = len(tables)
            metadata['table_data'] = tables

            full_text = '\n'.join([page['text'] for page in content])
            return full_text, metadata

        except Exception as e:
            logger.error(f"Error extracting PDF content: {e}")
            raise

    async def _extract_excel_content(self, file_path: Path, metadata: Dict[str, Any]) -> Tuple[str, Dict[str, Any]]:
        """Extract content from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = []
            all_text = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = {
                    'name': sheet_name,
                    'data': [],
                    'dimensions': f"{sheet.max_row}x{sheet.max_column}"
                }

                # Extract data from each row
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(row_data):  # Skip empty rows
                        sheet_data['data'].append(row_data)
                        all_text.extend(row_data)

                sheets_data.append(sheet_data)

            metadata.update({
                'sheet_count': len(workbook.sheetnames),
                'sheets': sheets_data,
                'max_row': max([s['dimensions'].split('x')[0] for s in sheets_data]) if sheets_data else 0,
                'max_col': max([s['dimensions'].split('x')[1] for s in sheets_data]) if sheets_data else 0
            })

            # Clean up text
            full_text = ' '.join([text for text in all_text if text.strip()])
            return full_text, metadata

        except Exception as e:
            logger.error(f"Error extracting Excel content: {e}")
            raise

    async def _extract_image_content(self, file_path: Path, metadata: Dict[str, Any]) -> Tuple[str, Dict[str, Any]]:
        """Extract text from image using OCR"""
        try:
            # Open image
            image = Image.open(file_path)

            # Update metadata
            metadata.update({
                'image_format': image.format,
                'image_mode': image.mode,
                'image_size': image.size,
                'dpi': image.info.get('dpi', (72, 72))
            })

            # Perform OCR
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')

            # Extract additional OCR data
            ocr_data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)

            metadata['ocr_confidence'] = np.mean([conf for conf in ocr_data['conf'] if conf > 0])
            metadata['word_count'] = len([word for word in ocr_data['text'] if word.strip()])

            return text.strip(), metadata

        except Exception as e:
            logger.error(f"Error extracting image content: {e}")
            raise

    async def _evaluate_quality(self, content: str, metadata: Dict[str, Any]) -> float:
        """Evaluate quality of extracted content"""
        quality_score = 0.0

        # Content length score (0-30 points)
        content_length = len(content)
        if content_length > 1000:
            quality_score += 30
        elif content_length > 500:
            quality_score += 20
        elif content_length > 100:
            quality_score += 10

        # Text density score (0-20 points)
        if content_length > 0:
            words = content.split()
            if metadata.get('file_size', 0) > 0:
                density = len(words) / metadata['file_size'] * 1000  # words per KB
                if density > 10:
                    quality_score += 20
                elif density > 5:
                    quality_score += 15
                elif density > 2:
                    quality_score += 10

        # Structure score (0-20 points)
        if '\n' in content:
            quality_score += 10
        if any(char in content for char in ['•', '-', '*', '1.', '2.']):
            quality_score += 10

        # Language and readability score (0-30 points)
        if self.language_detector:
            try:
                language_confidence = self.language_detector.predict([content])[0]
                quality_score += min(language_confidence * 30, 30)
            except:
                quality_score += 15  # Default score for unknown language
        else:
            quality_score += 15

        return min(quality_score, 100.0)

    async def _classify_document_type(self, content: str, metadata: Dict[str, Any]) -> str:
        """Classify document type based on content and metadata"""
        if self.document_classifier:
            try:
                features = self._extract_classification_features(content, metadata)
                prediction = self.document_classifier.predict([features])[0]
                return prediction
            except:
                pass

        # Fallback classification based on keywords
        content_lower = content.lower()

        if any(keyword in content_lower for keyword in ['报价', 'quote', '价格', 'price', 'cost']):
            return 'quotation'
        elif any(keyword in content_lower for keyword in ['合同', 'contract', '协议', 'agreement']):
            return 'contract'
        elif any(keyword in content_lower for keyword in ['发票', 'invoice', '账单', 'bill']):
            return 'invoice'
        elif any(keyword in content_lower for keyword in ['技术', 'technical', 'specification', '规格']):
            return 'technical_specification'
        elif any(keyword in content_lower for keyword in ['质量', 'quality', '检验', 'inspection']):
            return 'quality_document'
        elif any(keyword in content_lower for keyword in ['安全', 'safety', 'risk', '危险']):
            return 'safety_document'
        else:
            return 'general_document'

    def _extract_classification_features(self, content: str, metadata: Dict[str, Any]) -> np.ndarray:
        """Extract features for document classification"""
        features = []

        # Basic text features
        features.append(len(content))
        features.append(len(content.split()))
        features.append(len(set(content.split())))

        # Keyword features
        keywords = ['报价', '合同', '发票', '技术', '质量', '安全', 'quote', 'contract', 'invoice', 'technical']
        for keyword in keywords:
            features.append(content.lower().count(keyword))

        # Metadata features
        features.append(metadata.get('file_size', 0))
        features.append(metadata.get('page_count', 0))
        features.append(1 if metadata.get('tables_found', 0) > 0 else 0)

        return np.array(features)


class PriceAnalyzerAgent(EnhancedBaseAgent):
    """Advanced price analysis agent with ML-powered insights"""

    def __init__(self, orchestrator):
        super().__init__("price_analyzer", "Price Analyzer", orchestrator)
        self.capabilities.update({
            "price_analysis",
            "market_comparison",
            "trend_analysis",
            "price_optimization",
            "competitor_analysis",
            "cost_modeling",
            "price_prediction",
            "margin_analysis",
            "market_pricing_intelligence"
        })
        self.price_models = {}
        self.market_data = {}
        self.analysis_history = []

    async def initialize(self):
        """Initialize price analysis models"""
        await super().initialize()

        # Load price analysis models
        await self._load_price_models()

        # Load market data
        await self._load_market_data()

        logger.info("Price Analyzer Agent initialized with ML models")

    async def _load_price_models(self):
        """Load price prediction and analysis models"""
        try:
            self.price_predictor = joblib.load('models/price_predictor.pkl')
            self.price_classifier = joblib.load('models/price_classifier.pkl')
            self.anomaly_detector = IsolationForest(contamination=0.1)
        except FileNotFoundError:
            logger.warning("Price models not found, initializing new models")
            self.price_predictor = RandomForestRegressor(n_estimators=100)
            self.price_classifier = RandomForestClassifier(n_estimators=100)
            self.anomaly_detector = IsolationForest(contamination=0.1)

    async def _load_market_data(self):
        """Load historical market data"""
        try:
            conn = sqlite3.connect('knowledge_base.db')
            query = """
                SELECT fq.*, f.name as factory_name, f.location
                FROM factory_quotes fq
                JOIN factories f ON fq.factory_id = f.id
                ORDER BY fq.quote_date DESC
            """
            self.market_data = pd.read_sql_query(query, conn)
            conn.close()

            logger.info(f"Loaded {len(self.market_data)} market data points")
        except Exception as e:
            logger.error(f"Error loading market data: {e}")
            self.market_data = pd.DataFrame()

    async def _execute_task_logic(self, task: TaskRequest) -> Dict[str, Any]:
        """Execute price analysis task"""
        task_type = task.task_type
        parameters = task.parameters

        if task_type == "analyze_quote":
            return await self._analyze_price_quote(parameters)
        elif task_type == "market_comparison":
            return await self._compare_market_prices(parameters)
        elif task_type == "trend_analysis":
            return await self._analyze_price_trends(parameters)
        elif task_type == "price_prediction":
            return await self._predict_price(parameters)
        elif task_type == "competitor_analysis":
            return await self._analyze_competitors(parameters)
        elif task_type == "margin_optimization":
            return await self._optimize_margins(parameters)
        elif task_type == "detect_anomalies":
            return await self._detect_price_anomalies(parameters)
        else:
            raise ValueError(f"Unknown task type: {task_type}")

    async def _analyze_price_quote(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze a specific price quote"""
        quote_data = parameters.get('quote_data')
        if not quote_data:
            raise ValueError("Quote data is required")

        try:
            # Extract key features
            features = self._extract_price_features(quote_data)

            # Predict reasonable price range
            if not self.market_data.empty:
                predicted_range = self._predict_price_range(features)
            else:
                predicted_range = None

            # Market comparison
            market_comparison = await self._get_market_comparison(features)

            # Price competitiveness score
            competitiveness_score = self._calculate_competitiveness(
                quote_data.get('price', 0), market_comparison
            )

            # Risk analysis
            risk_analysis = self._analyze_price_risk(quote_data, market_comparison)

            # Recommendations
            recommendations = self._generate_price_recommendations(
                quote_data, market_comparison, competitiveness_score
            )

            return {
                'quote_analysis': {
                    'quote_price': quote_data.get('price'),
                    'predicted_range': predicted_range,
                    'competitiveness_score': competitiveness_score,
                    'market_position': self._determine_market_position(competitiveness_score),
                    'risk_analysis': risk_analysis,
                    'recommendations': recommendations
                },
                'market_comparison': market_comparison,
                'analysis_metadata': {
                    'analyzed_at': datetime.now().isoformat(),
                    'data_points_used': len(self.market_data),
                    'confidence_score': self._calculate_analysis_confidence(features, market_comparison)
                }
            }

        except Exception as e:
            logger.error(f"Error analyzing price quote: {e}")
            raise

    def _extract_price_features(self, quote_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract features for price analysis"""
        features = {}

        # Product features
        features['product_category'] = quote_data.get('product_category', '')
        features['material'] = quote_data.get('material', '')
        features['specification'] = quote_data.get('specification', '')
        features['quantity'] = float(quote_data.get('quantity', 1))
        features['unit'] = quote_data.get('unit', 'piece')

        # Price features
        features['price'] = float(quote_data.get('price', 0))
        features['unit_price'] = features['price'] / max(features['quantity'], 1)

        # Manufacturer features
        features['manufacturer'] = quote_data.get('manufacturer', '')
        features['location'] = quote_data.get('location', '')

        # Technical features
        features['complexity'] = self._assess_complexity(quote_data)
        features['quality_grade'] = quote_data.get('quality_grade', 'standard')

        # Time features
        features['quote_date'] = quote_data.get('quote_date', datetime.now().isoformat())

        return features

    def _assess_complexity(self, quote_data: Dict[str, Any]) -> str:
        """Assess product complexity"""
        complexity_score = 0

        # Check for complex specifications
        spec = quote_data.get('specification', '').lower()
        if any(term in spec for term in ['precision', 'tolerance', 'tight', '±0.01']):
            complexity_score += 3
        elif any(term in spec for term in ['standard', 'normal', '±0.1']):
            complexity_score += 1

        # Check for material complexity
        material = quote_data.get('material', '').lower()
        if any(term in material for term in ['titanium', 'inconel', 'stainless steel 316']):
            complexity_score += 2
        elif any(term in material for term in ['steel', 'aluminum']):
            complexity_score += 1

        # Check for process complexity
        if quote_data.get('special_processes'):
            complexity_score += len(quote_data['special_processes'])

        if complexity_score >= 5:
            return 'high'
        elif complexity_score >= 3:
            return 'medium'
        else:
            return 'low'

    async def _get_market_comparison(self, features: Dict[str, Any]) -> Dict[str, Any]:
        """Get market comparison data"""
        if self.market_data.empty:
            return {'status': 'no_data_available'}

        try:
            # Filter similar products
            similar_products = self.market_data[
                (self.market_data['product_category'] == features['product_category']) |
                (self.market_data['material'] == features['material'])
            ]

            if similar_products.empty:
                similar_products = self.market_data  # Use all data if no specific matches

            # Calculate statistics
            current_price = features['price']
            prices = similar_products['price'].dropna()

            comparison = {
                'total_comparisons': len(prices),
                'price_statistics': {
                    'mean': float(prices.mean()),
                    'median': float(prices.median()),
                    'std': float(prices.std()),
                    'min': float(prices.min()),
                    'max': float(prices.max()),
                    'percentile_25': float(prices.quantile(0.25)),
                    'percentile_75': float(prices.quantile(0.75))
                },
                'price_position': self._calculate_price_position(current_price, prices),
                'competitor_count': similar_products['factory_id'].nunique(),
                'data_date_range': {
                    'start': similar_products['quote_date'].min(),
                    'end': similar_products['quote_date'].max()
                }
            }

            return comparison

        except Exception as e:
            logger.error(f"Error in market comparison: {e}")
            return {'status': 'error', 'message': str(e)}

    def _calculate_price_position(self, current_price: float, market_prices: pd.Series) -> str:
        """Calculate price position in market"""
        percentile = (market_prices < current_price).mean()

        if percentile < 0.25:
            return 'low_price'
        elif percentile < 0.5:
            return 'below_average'
        elif percentile < 0.75:
            return 'above_average'
        else:
            return 'high_price'

    def _calculate_competitiveness(self, current_price: float, market_comparison: Dict[str, Any]) -> float:
        """Calculate price competitiveness score (0-100)"""
        if market_comparison.get('status') == 'no_data_available':
            return 50.0  # Neutral score

        try:
            stats = market_comparison['price_statistics']
            mean_price = stats['mean']
            std_price = stats['std']

            # Calculate z-score
            if std_price > 0:
                z_score = (current_price - mean_price) / std_price
            else:
                z_score = 0

            # Convert to competitiveness score (inverse of z-score)
            competitiveness_score = 50 - (z_score * 10)
            return max(0, min(100, competitiveness_score))

        except:
            return 50.0

    def _determine_market_position(self, competitiveness_score: float) -> str:
        """Determine market position based on competitiveness score"""
        if competitiveness_score >= 80:
            return 'highly_competitive'
        elif competitiveness_score >= 60:
            return 'competitive'
        elif competitiveness_score >= 40:
            return 'moderate'
        elif competitiveness_score >= 20:
            return 'expensive'
        else:
            return 'very_expensive'

    def _analyze_price_risk(self, quote_data: Dict[str, Any], market_comparison: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze price-related risks"""
        risks = []
        risk_level = 'low'

        current_price = quote_data.get('price', 0)

        # Check if price is significantly above market
        if market_comparison.get('price_statistics'):
            stats = market_comparison['price_statistics']
            if current_price > stats['percentile_75']:
                risks.append('Price significantly above market average')
                risk_level = 'medium'
            elif current_price > stats['max']:
                risks.append('Price exceeds maximum market price')
                risk_level = 'high'

        # Check supplier reliability
        manufacturer = quote_data.get('manufacturer', '')
        if not manufacturer or manufacturer.lower() in ['unknown', 'tbd']:
            risks.append('Unknown manufacturer')
            risk_level = 'medium' if risk_level == 'low' else risk_level

        # Check for missing specifications
        if not quote_data.get('specification'):
            risks.append('Missing product specifications')
            risk_level = 'medium' if risk_level == 'low' else risk_level

        return {
            'risk_level': risk_level,
            'identified_risks': risks,
            'risk_score': len(risks) * 10,
            'mitigation_suggestions': [
                'Verify supplier credentials',
                'Get multiple quotes',
                'Clarify specifications',
                'Check market references'
            ]
        }

    def _generate_price_recommendations(self, quote_data: Dict[str, Any],
                                       market_comparison: Dict[str, Any],
                                       competitiveness_score: float) -> List[str]:
        """Generate price recommendations"""
        recommendations = []

        if competitiveness_score < 40:
            recommendations.append('Consider negotiating for lower prices')
            recommendations.append('Request volume discounts')
            recommendations.append('Explore alternative suppliers')
        elif competitiveness_score > 80:
            recommendations.append('Current price is highly competitive')
            recommendations.append('Consider locking in long-term contracts')
            recommendations.append('Validate quality standards')

        # Market-specific recommendations
        if market_comparison.get('competitor_count', 0) < 3:
            recommendations.append('Limited competition in market - consider strategic partnerships')
        elif market_comparison.get('competitor_count', 0) > 10:
            recommendations.append('Highly competitive market - maintain price vigilance')

        # Product-specific recommendations
        complexity = self._assess_complexity(quote_data)
        if complexity == 'high':
            recommendations.append('Verify technical capabilities for complex specifications')
            recommendations.append('Consider quality assurance costs in price evaluation')

        return recommendations

    def _calculate_analysis_confidence(self, features: Dict[str, Any],
                                     market_comparison: Dict[str, Any]) -> float:
        """Calculate confidence score for the analysis"""
        confidence = 50.0  # Base confidence

        # Data availability confidence
        if market_comparison.get('total_comparisons', 0) > 10:
            confidence += 20
        elif market_comparison.get('total_comparisons', 0) > 5:
            confidence += 10

        # Feature completeness confidence
        required_features = ['product_category', 'material', 'specification', 'quantity']
        complete_features = sum(1 for feature in required_features if features.get(feature))
        confidence += (complete_features / len(required_features)) * 20

        # Market data freshness confidence
        if market_comparison.get('data_date_range'):
            # Assume recent data if date range includes last 30 days
            confidence += 10

        return min(100.0, confidence)


# Factory functions
def create_document_processor_agent(orchestrator) -> DocumentProcessorAgent:
    """Create a document processor agent"""
    return DocumentProcessorAgent(orchestrator)


def create_price_analyzer_agent(orchestrator) -> PriceAnalyzerAgent:
    """Create a price analyzer agent"""
    return PriceAnalyzerAgent(orchestrator)


# Usage example
if __name__ == "__main__":
    from multi_agent_orchestrator import MultiAgentOrchestrator

    async def test_specialized_agents():
        # Create orchestrator
        orchestrator = MultiAgentOrchestrator()
        await orchestrator.initialize()

        # Create specialized agents
        doc_agent = create_document_processor_agent(orchestrator)
        price_agent = create_price_analyzer_agent(orchestrator)

        # Register agents
        await orchestrator.register_agent(doc_agent)
        await orchestrator.register_agent(price_agent)

        # Test document processing
        doc_task = AgentTask(
            task_id="doc_test_001",
            task_type="process_document",
            parameters={'file_path': '/path/to/test.pdf'},
            priority=Priority.NORMAL
        )

        doc_result = await doc_agent.execute_task(doc_task)
        print(f"Document processing result: {doc_result.success}")

        # Test price analysis
        price_task = AgentTask(
            task_id="price_test_001",
            task_type="analyze_quote",
            parameters={
                'quote_data': {
                    'price': 1000,
                    'product_category': 'bolts',
                    'material': 'steel',
                    'quantity': 100,
                    'specification': 'M10x50'
                }
            },
            priority=Priority.NORMAL
        )

        price_result = await price_agent.execute_task(price_task)
        print(f"Price analysis result: {price_result.success}")

    asyncio.run(test_specialized_agents())